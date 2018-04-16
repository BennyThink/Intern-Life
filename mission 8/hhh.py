#!/usr/bin/python
# coding:utf-8

# Intern-Life - hhh.py
# 2018/4/16 13:49
# 

__author__ = 'Benny <benny@bennythink.com>'

import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill

data_raw = json.load(open('data.json'))
columns_raw = json.load(open('columns.json'))

filename = columns_raw['excel_name']
sheet_name = columns_raw['sheet_name']


def get_data(col, default):
    return [k.get(col) if k.get(col) != '' else default for k in data_raw]


def form_ds():
    new = {}
    for i in columns_raw['titles']:
        new[i] = {"label": columns_raw['titles'][i]['label'],
                  "value": get_data(columns_raw['titles'][i]['key'], columns_raw['titles'][i]['default'])}
    return new


def write_xls(f, _sheet_name, data):
    wb = Workbook()
    ws = wb.create_sheet()
    ws.title = _sheet_name
    # ws.column_dimensions['A'].width = 25
    # ws['B4'].fill = PatternFill(fill_type='solid', fgColor="5F9EA0")
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

    for j in data:
        length = []
        ws[j + '1'] = data[j]['label']
        ws[j + '1'].fill = PatternFill(fill_type='solid', fgColor="D3D3D3")
        for i in range(len(data[j]['value'])):
            length.append(len(data[j]['value'][i].encode('utf-8')) + 2)
            ws[j + str(i + 2)] = data[j]['value'][i]
        ws.column_dimensions[j].width = max(length)

    wb.save(f)


if __name__ == '__main__':
    form_data = form_ds()
    write_xls(filename, sheet_name, form_data)
